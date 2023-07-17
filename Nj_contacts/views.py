from django.shortcuts import render, HttpResponse, redirect

from . models import Contact, KCPE_collection_point, Kepsea_collection_point
from . forms import Contacts_Form, KCPE_collection_points_form, KEPSEA_collection_points_form
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from django.db.models import Q

# Create your views here.
def home(request):
    # total number of schools in the sub county
    school_count = Contact.objects.all().count # counts the total number of schools available
   
    # total number of PRIMARY schools
    primary_schools = Contact.objects.filter(school_category='PRIMARY').count
    public_primary_schools = Contact.objects.filter(type_of_school='PUBLIC', school_category='PRIMARY').count # counts the number of PUBLIC PRIMARY schools available
    private_primary_schools = Contact.objects.filter(type_of_school='PRIVATE', school_category='PRIMARY').count # counts the number of PRIVATE PRIMARY schools available

    # total number of SECONDARY schools
    secondary_schools = Contact.objects.filter(school_category='SECONDARY').count
    
    public_secondary_schools = Contact.objects.filter(type_of_school='PUBLIC', school_category='SECONDARY').count # counts the number of PUBLIC SECONDARY schools available
    private_secondary_schools = Contact.objects.filter(type_of_school='PRIVATE', school_category='SECONDARY').count # counts the number of PRIVATE SECONDARY schools available
   
    # total KCPE collection points
    KCPE_collection_point_count = KCPE_collection_point.objects.all().count

    KEPSEA_collection_point_count = Kepsea_collection_point.objects.all().count # total kepsea collection points

    # last updated item/ school added
    contact = Contact.objects.last()
    last_updated = contact.created_at    # gets the time at which the last school was added

    context = {
      'school_count' : school_count,
      'primary_schools': primary_schools,
      'public_primary_schools': public_primary_schools,
      'private_primary_schools' : private_primary_schools,
      'secondary_schools': secondary_schools,
      'public_secondary_schools': public_secondary_schools,
      'private_secondary_schools': private_secondary_schools,

      'last_updated' : last_updated,

      'KCPE_collection_point_count': KCPE_collection_point_count,

      'KEPSEA_collection_point_count': KEPSEA_collection_point_count,

    }

    return render(request, 'pages/Homepage.html', context=context)
    

@login_required(login_url='')   
# to display info on the webpage
def display_contacts(request):
    all_Contacts = Contact.objects.all()

    school_count = Contact.objects.all().count

    # performing searches
    if 'q' in request.GET:
        q=request.GET['q']

         # Filtering by school name or school code
        all_Contacts = Contact.objects.filter(Q(school_name__icontains=q) | Q(school_code__icontains=q))

    # all_Contacts = Contact.objects.all()

    context = {
      'all_Contacts': all_Contacts,
      'school_count' : school_count
    }

    return render(request, 'pages/contacts.html', context=context)

@login_required(login_url='')  
def show_Contact_Form(request):
    form = Contacts_Form()

    if request.method == 'POST':
        form = Contacts_Form(request.POST)

        # checking validity of the form b4 saving
        if form.is_valid():
            form.save() 
            messages.success(request, 'school added')

            return redirect('display_Contacts')
        
    context = {
        'form' : form
    }

    return render(request, 'pages/contacts_form.html', context=context)

@login_required(login_url='')  
# displays the school details
def school_details(request, pk):
    specific_School = Contact.objects.get(id=pk)

    context = {
       'specific_School': specific_School
    }

    return render(request, 'pages/school_details.html', context=context)

@login_required(login_url='')  
# function to delete selected school
def delete_school_record(request, pk):
    if request.user.is_authenticated:
        delete_school_record = Contact.objects.get(id=pk)
        delete_school_record.delete()
        messages.success(request, "School Deleted")
        return redirect('display_Contacts')
    else:
        messages.success(request, "You must be logged in")
        return redirect('')

@login_required(login_url='')     
def update_school_record(request, pk):
    if request.user.is_authenticated:
        current_record = Contact.objects.get(id=pk)
        form = Contacts_Form(request.POST or None, instance=current_record)
        if form.is_valid():
            form.save()
            messages.success(request, "Record Updated")
            return redirect('display_Contacts')
        return render(request, 'pages/contacts_form.html', {'form': form})
    else:
        messages.success(request, "You must be logged in")
        return redirect('home')

# to display the privacy policy page
def privacy_policy(request):
    return render(request, 'pages/privacy_policy.html')

# ==============================================================#
# Exporting data as excel
import openpyxl
from django.http import HttpResponse

@login_required(login_url='')  
def export_data_as_excel(request):
    data = Contact.objects.all()

    # Create a new workbook and get the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Get the fields dynamically from the model
    fields = Contact._meta.fields

    # Write the header row with field names
    header_row = [field.name for field in fields]
    for col_num, field_name in enumerate(header_row, start=1):
        sheet.cell(row=1, column=col_num, value=field_name)

    # Write the data rows
    for row_num, item in enumerate(data, start=2):
        for col_num, field in enumerate(fields, start=1):
            field_value = getattr(item, field.name)
            sheet.cell(row=row_num, column=col_num, value=field_value)

    # Set the response headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


# count the number of schools
def dashboard(request):
    school_count = Contact.objects.count()
    context = {'school_count': school_count}
    return render(request, 'dashboard.html', context)

# ====================////// ENROLMENT NUMBER ===============================================



# ====================////// EXAMINATION COLLECTION POINTS ===============================================
@login_required(login_url='')   
# to display info on KCPE collection points
def display_kcpe_collection_points(request):
    all_collection_points = KCPE_collection_point.objects.all()

    collection_points_count = KCPE_collection_point.objects.all().count

    # performing searches
    if 'q' in request.GET:
        q=request.GET['q']

         # Filtering by collection point by name or school code
        all_collection_points = KCPE_collection_point.objects.filter(Q(school_name__icontains=q) | Q(school_code__icontains=q))

    context = {
      'all_collection_points': all_collection_points,
      'collection_points_count' : collection_points_count
    }

    return render(request, 'pages/collection_points.html', context=context)

# to display the KCPE collection form
@login_required(login_url='') 
def show_KCPE_collection_form(request):
    form = KCPE_collection_points_form
    
    if request.method == 'POST':
        form = KCPE_collection_points_form(request.POST)

        # checking validity of the form b4 saving
        if form.is_valid():
            form.save() 
            messages.success(request, 'collection point added')

            return redirect('display_kcpe_collection_points')
        
    context = {
        'form' : form
    }

    return render(request, 'pages/KCPE_collectionpoints_form.html', context=context)

# to update KCPE school collection
@login_required(login_url='')     
def update_KCPE_collection_point(request, pk):
    if request.user.is_authenticated:
        current_KCPE_collection_point = KCPE_collection_point.objects.get(id=pk)
        form = KCPE_collection_points_form(request.POST or None, instance=current_KCPE_collection_point)
        if form.is_valid():
            form.save()
            messages.success(request, "Record Updated")
            return redirect('display_kcpe_collection_points')
        return render(request, 'pages/KCPE_collectionpoints_form.html', {'form': form})
    else:
        messages.success(request, "You must be logged in")
        return redirect('')
    

# display KCPE collection point in details
@login_required(login_url='')  
def display_detailed_KCPE_collection_point(request, pk):
    specific_collection_point = KCPE_collection_point.objects.get(id=pk)

    context = {
       'specific_collection_point': specific_collection_point
    }

    return render(request, 'pages/kcpe_collection_point_details.html', context=context)


# =======================================================
from django.views.decorators.http import require_http_methods

require_http_methods(['DELETE'])
# deleting a collection point
def delete_KCPE_collection_point(request, pk):
    delete_collection_point = KCPE_collection_point.objects.get(id=pk)
    delete_collection_point.delete()

    return redirect('display_kcpe_collection_points')


# ====================//// KEPSEA COLLECTION POINT VIEWS =================
# display the KEPSEA collection points
# @login_required(login_url='') 
def display_kepsea_collection_points(request):
    all_collection_points = Kepsea_collection_point.objects.all()

    collection_points_count = Kepsea_collection_point.objects.all().count

    # performing searches
    if 'q' in request.GET:
        q=request.GET['q']

         # Filtering by collection point by name or school code
        all_collection_points = Kepsea_collection_point.objects.filter(Q(school_name__icontains=q) | Q(school_code__icontains=q))

    context = {
      'all_collection_points': all_collection_points,
      'collection_points_count' : collection_points_count
    }

    return render(request, 'pages/kepsea/display_kepsea_collection_points.html', context=context)

# display the KEPSEA collection form
# @login_required(login_url='') 
def show_KEPSEA_collection_form(request):
    form = KEPSEA_collection_points_form
    
    if request.method == 'POST':
        form = KEPSEA_collection_points_form(request.POST)

        # checking validity of the form b4 saving
        if form.is_valid():
            form.save() 
            messages.success(request, 'collection point added')

            return redirect('display_kepsea_collection_points')
        
    context = {
        'form' : form
    }

    return render(request, 'pages/kepsea/collection_form.html', context=context)


# ==================== END OF KEPSEA COLLECTION POINT VIEWS ==============



# ====================////// SEARCH OPTION USING AJAX FOR COLLECTIION POINTS /////==========

# search for collection points
from django.http import JsonResponse

# for users who are not logged in can search for their school
# === for KCPE
def search_school(request):
    query = request.GET.get('query', '')

    schools = KCPE_collection_point.objects.filter(
        Q(school_code__icontains=query) |   # searches by school_code
        Q(school_name__icontains=query)     # searches by school name
    )

    if schools.exists():
        school = schools.first()
        response = {
            'school_name': school.school_name,
            'entry': school.entry,
            'collection_point': school.collection_point
        }
    else:
        response = {'school_name': None, 'entry': None, 'collection_point': None}

    return JsonResponse(response)


# === for KEPSEA
def search_kepsea(request):
    query = request.GET.get('query', '')

    schools = Kepsea_collection_point.objects.filter(
        Q(school_code__icontains=query) |   # searches by school_code
        Q(school_name__icontains=query)     # searches by school name
    )

    if schools.exists():
        school = schools.first()
        response = {
            'school_name': school.school_name,
            'entry': school.entry,
            'collection_point': school.collection_point
        }
    else:
        response = {'school_name': None, 'entry': None, 'collection_point': None}

    return JsonResponse(response)



    # query = request.GET.get('query', '')

    # try:
    #     # Searching by school code
    #     school = KCPE_collection_point.objects.get(school_code=query)
    # except KCPE_collection_point.DoesNotExist:
    #     try:
    #         # Searching by school name
    #         school = KCPE_collection_point.objects.get(school_name=query)
    #     except KCPE_collection_point.DoesNotExist:
    #         school = None

    # if school:
    #     response = {
    #         'school_name': school.school_name,
    #         'entry': school.entry,
    #         'collection_point': school.collection_point
    #     }
    # else:
    #     response = {'school_name': None, 'entry': None, 'collection_point': None}

    # return JsonResponse(response)
# =============== END OF SERACH OPTION ======================================


    









    
    






